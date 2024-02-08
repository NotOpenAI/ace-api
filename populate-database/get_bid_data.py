import openpyxl
import random
import psycopg2

work_book = openpyxl.load_workbook("ace-api/populate-database/AI_Sample_Data.xlsx")

workbook_contents = work_book.active

column_headers = list()
bids = list()

# Grab all the column headers
for i in range(1, workbook_contents.max_column + 1):
    column_headers.append(workbook_contents.cell(row=1, column=i).internal_value)

# iterate through excel and store bid data
for i in range(2, workbook_contents.max_row + 1):
    bid_data = dict()
    print("\n")
    for j in range(1, workbook_contents.max_column + 1):
        cell_value = str(workbook_contents.cell(row=i, column=j).value)
        if "RAND" in cell_value:
            if "(0,4)" in cell_value:
                cell_value = str(random.randint(0, 4))
            elif "(0,3)" in cell_value:
                cell_value = str(random.randint(0, 3))
        bid_data[column_headers[j - 1]] = cell_value
    bids.append(bid_data)
    print(str(column_headers[j - 1]) + ": " + str(cell_value))

for bid in bids:
    for key in bid.keys():
        print(key + ": " + bid[key])
